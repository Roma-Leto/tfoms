# region Imports
import logging
import re, os
import json
import tempfile

from django.core.files import File
# from msilib.schema import Font

from openpyxl.styles import Font
import openpyxl.cell.cell
from django.db.models.expressions import result, F
from openpyxl import Workbook
from celery import shared_task
from pathlib import Path
from openpyxl import load_workbook
from django.shortcuts import redirect
from datetime import datetime
from django.db import IntegrityError, connection
from openpyxl.styles import Alignment
from openpyxl.styles.builtins import title
from openpyxl.utils import get_column_letter, column_index_from_string

import invoice.validators
# from utilities import timer
from x_tfoms_project import settings
from invoice.models import InvoiceDNRDetails, InvoiceAttachment, InvoiceInvoiceJobs, InvoiceInvoiceJobSteps, FileUpload

# endregion Imports
logger = logging.getLogger(__name__)

# Build paths inside the project like this: BASE_DIR / 'subdir'.
BASE_DIR = Path(__file__).resolve().parent.parent

def parse_second_sheet(data_excel):
    """
    Функция парсинга excel-файла
    :param data_excel: кортеж данных, извлечённых со страницы документа
    :return: словарь result
    """
    result = dict()
    try:
        # Проверка данных перед обработкой
        # print("data_excel", data_excel)
        # if len(data_excel) > 0 and len(data_excel[0]) > 14:
        if True:
            result['conditions_of_medical_care'] = data_excel[0]
            # logger.info(f"Кода вида и условий оказания медицинской помощи "
            #             f"{result['conditions_of_medical_care']}")
            result['usl_ok'] = 5 - int(data_excel[0][0])
            result['mocod'] = data_excel[2]
            result['tip'] = data_excel[3]
            result['patients_name'] = data_excel[1]
            # logger.info(f"ФИО {result['patients_name']}")
            result['birthday'] = data_excel[4]
            # logger.info(f"Дата рождения {result['birthday']}")
            result['policy_number'] = data_excel[5]
            # logger.info(f"Номер полиса(ЕНП) {result['policy_number']}")
            delimiters = r'[()-]'  # Символы-разделители
            result['medical_care_profile_code'] = \
                find_medical_doctor_code(re.split(delimiters, data_excel[7]))[0][0]
            # logger.info(f"Код профиля медицинской помощи "
            #             f"{result['medical_care_profile_code']}")
            result['doctors_specialty_code'] = \
                find_medical_doctor_code(re.split(delimiters, data_excel[7]))[0][1]
            # logger.info(f"Код специальности врача "
            #             f"{result['doctors_specialty_code']}")

            result['medical_care_profile_name'] = \
                find_medical_doctor_code(re.split(delimiters, data_excel[7]))[1][0]
            # logger.info(f"Код профиля медицинской помощи "
            #             f"{result['medical_care_profile_code']}")
            result['doctors_specialty_name'] = \
                find_medical_doctor_code(re.split(delimiters, data_excel[7]))[1][1]
            # logger.info(f"Код специальности врача "
            #             f"{result['doctors_specialty_code']}")
            result['subj_n'] = data_excel[6]
            result['diagnosis'] = data_excel[8]
            # logger.info(f"Диагноз {result['diagnosis']}")
            result['start_date_of_treatment'] = data_excel[9]
            # logger.info(f"Дата начала лечения {result['start_date_of_treatment']}")
            result['end_date_of_treatment'] = data_excel[10]
            # logger.info(f"Дата окончания лечения {result['end_date_of_treatment']}")
            result['treatment_result_code'] = \
                re.split(delimiters, data_excel[11])[1]
            # logger.info(f"Дата окончания лечения {result['treatment_result_code']}")
            result['treatment_result_name'] = \
                re.split(delimiters, data_excel[11])[2]
            # logger.info(f"Дата окончания лечения {result['treatment_result_name']}")
            result['volume_of_medical_care'] = data_excel[12]
            # logger.info(
            # f"Объёма медицинской помощи {result['volume_of_medical_care']}")
            result['tariff'] = data_excel[13]
            # logger.info(f"Тариф {result['tariff']}")
            result['expenses'] = data_excel[14]
            # logger.info(f"Расходы {result['expenses']}")



    except IndexError as e:
        logger.error(f"Ошибка при обработке данных: {e}")

    logger.info(f"Result: {result}")

    return result


def parse_second_sheet_lnr(data_excel):
    """
    Функция парсинга excel-файла
    :param data_excel: кортеж данных, извлечённых со страницы документа
    :return: словарь result
    """
    result = dict()
    print("parse_second_sheet, data_excel", data_excel)
    try:
        # Проверка данных перед обработкой
        result['conditions_of_medical_care'] = data_excel[0]
        logger.info(f"Кода вида и условий оказания медицинской помощи "
                    f"{result['conditions_of_medical_care']}")
        result['usl_ok'] = 5 - int(data_excel[0][0])
        logger.info(f"usl_ok: {result['usl_ok']}")
        result['mocod'] = data_excel[2]
        logger.info(f"mocod: {result['mocod']}")
        result['tip'] = data_excel[3]
        logger.info(f"tip: {result['tip']}")
        result['patients_name'] = data_excel[1]
        logger.info(f"ФИО {result['patients_name']}")
        result['birthday'] = data_excel[4]
        logger.info(f"Дата рождения {result['birthday']}")
        result['policy_number'] = data_excel[5]
        logger.info(f"Номер полиса(ЕНП) {result['policy_number']}")
        delimiters = r'[()-]'  # Символы-разделители
        result['medical_care_profile_code'] = \
            find_medical_doctor_code(re.split(delimiters, data_excel[7]))[0][0]
        logger.info(f"Код профиля медицинской помощи "
                    f"{result['medical_care_profile_code']}")
        result['doctors_specialty_code'] = \
            find_medical_doctor_code(re.split(delimiters, data_excel[7]))[0][1]
        logger.info(f"Код специальности врача "
                    f"{result['doctors_specialty_code']}")

        result['medical_care_profile_name'] = \
            find_medical_doctor_code(re.split(delimiters, data_excel[7]))[1][0]
        logger.info(f"Код профиля медицинской помощи "
                    f"{result['medical_care_profile_code']}")
        result['doctors_specialty_name'] = \
            find_medical_doctor_code(re.split(delimiters, data_excel[7]))[1][1]
        logger.info(f"Код специальности врача "
                    f"{result['doctors_specialty_code']}")
        result['subj_n'] = data_excel[6]
        result['diagnosis'] = data_excel[8]
        logger.info(f"Диагноз {result['diagnosis']}")
        result['start_date_of_treatment'] = data_excel[9]
        logger.info(f"Дата начала лечения {result['start_date_of_treatment']}")
        result['end_date_of_treatment'] = data_excel[10]
        logger.info(f"Дата окончания лечения {result['end_date_of_treatment']}")
        result['treatment_result_code'] = \
            re.split(delimiters, data_excel[11])[1]
        logger.info(f"Дата окончания лечения {result['treatment_result_code']}")
        result['treatment_result_name'] = \
            re.split(delimiters, data_excel[11])[2]
        logger.info(f"Дата окончания лечения {result['treatment_result_name']}")
        result['volume_of_medical_care'] = data_excel[12]
        logger.info(
        f"Объёма медицинской помощи {result['volume_of_medical_care']}")
        result['tariff'] = data_excel[13]
        logger.info(f"Тариф {result['tariff']}")
        result['expenses'] = data_excel[14]
        logger.info(f"Расходы {result['expenses']}")



    except IndexError as e:
        logger.error(f"Ошибка при обработке данных: {e}")

    logger.info(f"Result: {result}")

    return result


def convert_date(report_date_str) -> datetime.date:
    """
    Преобразование формата даты
    :param report_date_str: Строка в формате dd.mm.yyyy
    :return report_date: Дата в формате YYYY-MM-DD
    """
    try:
        # Преобразуем строку в объект datetime
        report_date = datetime.strptime(report_date_str, "%d.%m.%Y").date()
        return report_date
    except ValueError as e:
        print(f"Неверный формат даты: {e}")
        return False


def find_medical_doctor_code(lst: list):
    """
    Ищет в профиле оказания медицинской помощи - специальности врача цифровые коды
    :param lst:
    :return:
    """
    numbers = []
    names = []
    for item in lst:
        try:
            num = int(item)  # Сначала попробуем преобразовать в целое число
            numbers.append(num)
        except ValueError:
            try:
                num = float(item)  # Затем попробуем преобразовать в вещественное число
                numbers.append(num)
            except ValueError:
                if len(item) > 2:
                    names.append(item)
                continue  # Если не получилось ни то, ни другое, пропускаем элемент

        if len(numbers) >= 2 and len(names) >= 2:
            break  # Останавливаемся, как только нашли два числа
    return numbers, names


def call_procedure(ext_id):
    logger.info("call_procedure start")
    with connection.cursor() as cursor:
        # Вызов хранимой процедуры с параметром
        cursor.execute("EXEC dbo.check_invoice_flk @ext_id = %s", [ext_id])
        # row = cursor.fetchone()
        # logger.info("call_procedure end", row)
        logger.info("call_procedure end")
    return 0


def get_errors(id):
    """Функция извлекает данные из поля errore_list таблицы invoice_errors"""
    with connection.cursor() as cursor:
        query = """
            SELECT *
            FROM invoice_errors
            WHERE attachment_id = %s
        """

        params = [id]

        cursor.execute(query, params)

        data = cursor.fetchall()

        return data

def create_report(ext_id):
    logger.info("Создание итогового отчёта...")

    # Получаем объект InvoiceDNRDetails по идентификатору ext_id
    try:
        invoice_dnr_details = InvoiceDNRDetails.objects.get(id=ext_id)
    except InvoiceDNRDetails.DoesNotExist:
        raise Exception(f"Не найден объект InvoiceDNRDetails с id={ext_id}")
    # Теперь получаем связанный объект FileUpload
    try:
        file_upload = FileUpload.objects.get(parent=invoice_dnr_details.id)
    except FileUpload.DoesNotExist:
        raise Exception("Запись не найдена.")
    # Создаем временный файл
    temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx")

    # region Создание документа в памяти
    # Создание рабочей книги
    wb = Workbook()

    # Создание дополнительных листов
    ws2 = wb.create_sheet("Справочник")
    ws3 = wb.create_sheet("Итого")

    # Получить активный лист
    ws = wb.active
    # Переименовать лист по умолчанию
    ws.title = 'Сводная таблица'
    # endregion Создание документа в памяти

    # region Формирование шапки документа. Страница 1
    # Объединяем ячейки B2:D2 (это будет объединённая ячейка)
    ws.merge_cells('A1:N1')

    # Записываем данные в объединённую ячейку
    ws['A1'] = 'Коды ошибок'

    # Центруем текст по горизонтали и вертикали
    alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].alignment = alignment

    # Шапка первой страницы документа
    top_data = [
        # [], # Использовать в случае отсутствия объединения строк выше
        ["ЗЛ не идентифицировано в ЕРЗ",
        "Уникальный идентификатор МО отсутствует в ФРМО",
        "Диагноз отсутствует в справочнике МКБ10",
        "Диагноз должен быть указан с точностью до подрубрики",
        "Код профиля МП отсутствует в классификаторе V002",
        "Некорректное сочетание Пол ЗЛ + Диагноз",
        "МКБ не входит в справочник диагнозов, оплачиваемых из средств ОМС",
        "Код результата обращения отсутствует в справочнике V009",
        "Код специальности отсутствует в классификаторе V021",
        "Некорректное сочетание Профиль МП + Возраст ЗЛ",
        "Некорректное сочетание Специальность + Возраст ЗЛ",
        "Некорректное сочетание Профиль МП + Пол ЗЛ",
        "Некорректное сочетание Специальность + Условия оказания МП",
        "Приоритетная ошибка",
        "№ п/п",
        "Фамилия Имя Отчество",
        "Номер в сводном реестре мед. организаций",
        "Дата рождения застрахованного лица",
        "Номер полиса обязательного мед. страхования ЗЛ",
        "Код профиль медицинской помощи",
        "Профиль оказания медицинской помощи",
        "Код специальности врача",
        "Специальность врача",
        "Диагноз (МКБ-10) застрахованного лица",
        "Дата начала лечения застрахованного лица",
        "Дата окончания лечения застрахованного лица",
        "Код результата лечения",
        "Результат лечения застрахованного лица",
        "Объемы предоставленной медицинской помощи",
        "Средний норматив фин. затрат на ед. объема мед. помощи (рублей)",
        "Расходы на оказание медицинской помощи (рублей)"
         ]
    ]

    for row in top_data:
        ws.append(row)
    # Установка высоты строки
    ws.row_dimensions[2].height = 80  # Высота первой строки равна 60 пунктов

    # Определяем желаемую ширину столбцов
    desired_width = 15  # Ширина в единицах (примерно соответствует ширине символа)

    # Итерируемся по столбцам от A до AR
    start_column = 'A'
    end_column = 'AT'

    # Переводим буквенные обозначения в числовые индексы
    start_col_idx = column_index_from_string(start_column)
    end_col_idx = column_index_from_string(end_column)

    # Итерируем по числовым индексам столбцов
    for col_num in range(start_col_idx, end_col_idx + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = desired_width


    # Включаем перенос текста в ячейке
    for row in ws["A1:AR2"]:
        for cell in row:
            cell.alignment = Alignment(wrapText=True)

    # Жирный шрифт
    bolt_font = Font(bold=True)
    for row in ws["O1:AR2"]:
        for cell in row:
            cell.font = bolt_font

    ws.column_dimensions['O'].width = 7
    ws.column_dimensions['P'].width = 45
    ws.column_dimensions['S'].width = 30
    ws.column_dimensions['N'].width = 25
    ws.column_dimensions['U'].width = 35
    ws.column_dimensions['W'].width = 35
    ws.column_dimensions['AB'].width = 60

    # endregion Формирование шапки документа

    # region Формирование строки-записи документа. Страница 1

    # Получаем всех ЗЛ из документа (по ext_id)
    policyholder = InvoiceAttachment.objects.filter(ext_id=ext_id)

    # Начальная строка для вставки данных в отчёт
    starting_row = 3

    # Перебираем все ЗЛ
    for item in policyholder:
        id_item = item.id

        string_result = [
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            item.row_id,
            item.fio,
            item.mocod,
            item.dr,
            item.enp,
            item.profil_id,
            item.profil_n,
            item.spec_id,
            item.spec_n,
            item.dz,
            item.date1,
            item.date2,
            item.rslt_id,
            item.rslt_n,
            item.cnt_usl,
            item.tarif,
            item.sum_usl
        ]

        # Обработка данных об ошибках
        result_errors = get_errors(id_item)
        if result_errors:
            top_error = result_errors[0][3]
            other_errors = result_errors[0][2].strip('{[]}').split(',')
            string_result[13] = top_error
            for error in other_errors:
                string_result[int(error)-1] = '1'

        # print(string_result)



        # Цикл для вставки данных построчно
        for col_num, value in enumerate(string_result, start=1):
            ws.cell(row=starting_row, column=col_num).value = value

        starting_row += 1
    # endregion Формирование строки-записи документа. Страница 1

    # region Формирование справочника документа. Страница 2

    title_ws2 = [[
        '№',
        'Наименование ошибки',
        'Серьёзность'
    ]]

    about_errors =[
        ['1', 'не идентифицирован', 220],
        ['2', 'код МО неверен', 230],
        ['3', 'нет диагноза', 210],
        ['4', 'диагноз не точен', 10],
        ['5', 'неверный профиль', 170],
        ['6', 'неверно пол + диагноз', 180],
        ['7', 'диагноз не входит в ОМС', 200],
        ['8', 'код результата не верен', 20],
        ['9', 'неверная специальность', 160],
        ['10', 'неверный профиль + возраст', 80],
        ['11', 'неверная специальность + возраст', 70],
        ['12', 'неверный профиль + пол', 150],
        ['13', 'неверная специальность + усл.ок.', 140]
    ]

    # Формирование шапки таблицы второго листа
    for row in title_ws2:
        ws2.append(row)

    for row_index, row_data in enumerate(about_errors, start=2):
        for col_index, col_value in enumerate(row_data, start=1):
            ws2.cell(row=row_index, column=col_index).value = col_value

    # Определяем желаемую ширину столбца
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 17

    # Текст шапки жирным

    ws2['A1'].font = bolt_font
    ws2['B1'].font = bolt_font
    ws2['C1'].font = bolt_font

    # endregion Формирование справочника документа. Страница 2

    # region Формирование итоговой таблицы сумм. Страница 3

    # Определяем желаемую ширину столбца
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 20

    ws3['A1'].font = bolt_font
    ws3['A1'] = 'Номер счёта'
    ws3['B1'].font = bolt_font
    ws3['B1'] = 'Дата счёта'
    ws3['C1'].font = bolt_font
    ws3['C1'] = 'Сумма счёта'

    doc = InvoiceDNRDetails.objects.get(id=ext_id)

    ws3['A2'] = doc.invoice_number
    ws3['B2'] = doc.date_of_reporting_period
    ws3['C2'] = doc.total_amount

    headers_ws3 = ['Условия оказания', 'Количество', 'Сумма']

    for col, header in enumerate(headers_ws3, start=1):
        ws3.cell(row=4, column=col).value = header
        ws3.cell(row=4, column=col).font = bolt_font

    from django.db.models import Sum

    sum_by_category = InvoiceAttachment.objects.values('usl_ok').annotate(total_usl=Sum('cnt_usl'), total_sum=Sum(F('cnt_usl') * F('tarif')))



    sum_result = []
    for x in sum_by_category:
        sum_step = []
        for y in x.values():
            sum_step.append(y)
        sum_result.append(sum_step)
    print(sum_result)

    for row_index, row_data in enumerate(sum_result, start=5):
        for col_index, col_value in enumerate(row_data, start=1):
            ws3.cell(row=row_index, column=col_index).value = col_value

    # endregion Формирование итоговой таблицы сумм. Страница


    # region Сохранение в файл.
    # Путь к директории
    directory = "./media/results/"
    # file_path = os.path.join(directory, f"report_{item.invoice_number}.xlsx")
    # file_path = os.path.join(directory, f"report_test.xlsx")
    logger.info("Проверка STATUS")



    status4 = InvoiceInvoiceJobs.objects.get(ext_id=ext_id, step_id=4)
    print(status4)
    if status4.ready:
        file_path = "report_test.xlsx"

        # TODO: Добавить обработчики других форматов
        # Сохраняем файл Excel во временный файл
        wb.save(temp_file.name)
        # Определяем новое имя файла
        new_filename = f'result_{os.path.basename(file_upload.file.name)}'

        # Обновляем поле result_file
        with open(temp_file.name, 'rb') as file_data:
            file_upload.result_file.save(new_filename, File(file_data), save=False)

        # Сохраняем изменения
        file_upload.save()
        # wb.save(file_path)
        # endregion Сохранение в файл.
        step5 = InvoiceInvoiceJobs.objects.get(ext_id=ext_id, step_id=5)
        step5.ready = True
        step5.status = "Выполнено"
        step5.save()

    f = InvoiceDNRDetails.objects.get(id=ext_id)
    # f.file_name
    logger.info("Создание итогового отчёта - ОК")


@shared_task
def celery_save_second_sheet(invoice_number):
    """
    Парсинг и сохранения данных второго листа отчёта
    :return:
    """
    # region Поиск и загрузка файла счёта в память
    item = InvoiceDNRDetails.objects.get(invoice_number=invoice_number)
    filename = item.file_name.replace(' — ',
                                      '__')  # замена длинного тире на обычный дефис
    filename = item.file_name.replace(' ',
                                      '_')  # замена длинного тире на обычный дефис
    file_path = os.path.join(settings.MEDIA_ROOT, 'uploads', filename)
    # endregion Поиск и загрузка файла счёта в память

    # region Проверка открытия файла отчёта
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл не найден: {file_path}")
    try:
        workbook = load_workbook(file_path, data_only=True)
    except Exception as e:
        logger.info(f"Произошла ошибка при открытии файла: {e}")
    # endregion Проверка открытия файла отчёта

    # region Формируем список листов
    # Загрузка Excel-файла с помощью openpyxl
    # workbook = load_workbook(file_path, data_only=True)
    sheet_list = list()
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]  # Получаем лист по имени
        sheet_list.append(sheet)
        logger.info(f'Название листа: {sheet_name}')  # Выводим имя листа
    # endregion Формируем список листов

    # region Сохраняем каждую строку данных в базу данных
    data_excel = list()  # Создаём список для строк документа
    # Пропустим первые три строки
    start_row_index = 3  # Начинаем с 4-й строки (индексация с нуля)
    for row in sheet_list[1].iter_rows(min_row=start_row_index,
                                       values_only=True):
        print("ROW", row)
        if ( # Записью о ЗЛ считаются те строки в которых:
                # all(element is not None for element in row[0:10])  # нет пустых ячеек в первых 9 столбцах
                any(isinstance(element, datetime) for element in row[0:10] )  # Есть ли ячейки с датами
                and len(str(row[0])) >= 3  # в первой ячейке текст больше 3-х символов
        ):
            invoice.validators.validate_tuple(row)
            data_excel.append(row)

    # Установка флага записи в БД, если не установлен ранее
    if not InvoiceInvoiceJobs.objects.filter(ext_id=item.id).exists():
        step_save_ppl = InvoiceInvoiceJobs.objects.create(
            ext_id=item.id,
            step_id=1,
            ready=0,
            status="Выполняется"
        )

    # Извлечение данных по каждому пациенту в БД
    count = 0
    for pers in data_excel:
        # Извлекаем данные из ячеек документа и формируем словарь
        if item.code_fund.name == "ЛНР":
            clear_data = parse_second_sheet_lnr(pers)
        elif item.code_fund.name == "ДНР":
            clear_data = parse_second_sheet(pers)
        count += 1
        # Запись в БД
        try:
            InvoiceAttachment.objects.create(
                ext_id=InvoiceDNRDetails.objects.latest('id').id,
                usl_ok=clear_data['usl_ok'],
                mocod=clear_data['mocod'],
                tip=clear_data['tip'],
                row_id=clear_data['conditions_of_medical_care'],
                fio=clear_data['patients_name'],
                dr=convert_date(clear_data['birthday']),
                enp=int(clear_data['policy_number']),
                profil_id=clear_data['medical_care_profile_code'],
                spec_id=clear_data['doctors_specialty_code'],
                profil_n=clear_data['medical_care_profile_name'],
                spec_n=clear_data['doctors_specialty_name'],
                subj_n=clear_data['subj_n'],
                dz=clear_data['diagnosis'],
                date1=convert_date(
                    clear_data['start_date_of_treatment']),
                date2=convert_date(
                    clear_data['end_date_of_treatment']),
                rslt_id=clear_data['treatment_result_code'],
                rslt_n=clear_data['treatment_result_name'],
                cnt_usl=clear_data['volume_of_medical_care'],
                tarif=clear_data['tariff'],
                sum_usl=clear_data['expenses']
            )
        except IntegrityError as e:
            logger.info(f"Запись с такими параметрами уже существует. {e}")
    InvoiceInvoiceJobs.objects.update(ready=1, status="Выполнено")
    # step_save_ppl.ready = 1
    # step_save_ppl.status = "Выполнено"
    # step_save_ppl.save()

    # Вызов процедуры
    call_procedure(item.id)

    # Формирование результирующего отчёта
    create_report(item.id)
    return 0
