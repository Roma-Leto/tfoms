import logging
import os
import re

from openpyxl import load_workbook
from django.contrib.auth.views import LoginView
from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db import IntegrityError

from x_tfoms_project.celery import debug_task
from invoice.forms import UploadFileForm
from invoice.models import InvoiceDNRDetails, RegisterTerritorial, FileUpload
from invoice.views import parse_first_sheet, mouth_converter
from invoice.tasks import convert_date
logger = logging.getLogger(__name__)
# TODO: обработать ситуацию отсутствия файлов для скачивания и сокрытия ссылок

class TLoginView(LoginView):
    template_name = 'registration/login.html'


def clean_filename(filename):
    """
    Очищает имя файла от специальных символов.
    """
    # Убираем все недопустимые символы
    filename = re.sub(r'[\\/*?:"<>|— -]', '', filename)
    return filename

@login_required
def profile(request):
    logger.info("func profile")
    # Вывод результатов
    message = ''
    if request.method == 'POST':
        logger.info("func profile, method POST")
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            # Загрузка Excel-файла с помощью openpyxl
            workbook = load_workbook(file, data_only=True)
            # Получаем первый лист
            first_sheet_name = workbook.sheetnames[0]
            first_sheet = workbook[first_sheet_name]

            # Логируем название первого листа
            logger.info(f'Название листа: {first_sheet_name}')

            # Извлекаем данные из первого листа
            data_excel = [row for row in first_sheet.iter_rows(values_only=True)]

            # Извлекаем данные из ячеек документа и формируем словарь
            clear_data = parse_first_sheet(data_excel, file)
            code_from_register = RegisterTerritorial.objects.get(
                code=clear_data['code_fund'])
            # Создание записи первой страницы в БД
            try:
                inv_object = InvoiceDNRDetails.objects.create(
                    file_name=file,
                    # Передаём месяц в виде числа используя функцию конвертации
                    mouth_of_invoice_receipt=mouth_converter(
                        clear_data['mouth_of_invoice_receipt']),
                    year_of_invoice_receipt=clear_data['year_of_invoice_receipt'],
                    # Преобразование даты в формат YYYY-MM-DD
                    date_of_reporting_period=convert_date(
                        clear_data['date_of_reporting_period']),
                    code_fund=code_from_register,
                    invoice_number=clear_data['invoice_number'],
                    total_amount=clear_data['total_amount'],
                    # ext_id=clear_data['ext_id']
                )
                # Сохранение файла под номером счёта
                logger.info("func profile. Сохранение данных первой страницы - ОК")
            except IntegrityError as e:
                inv_object = InvoiceDNRDetails.objects.get(invoice_number=
                                                           clear_data['invoice_number'])
                logger.error(f"Ошибка: {e}")

            item = inv_object
            # Создаем объект модели и сохраняем файл
            # Сохранение записи о файле
            try:
                # Очищаем имя файла для сохранения на диск
                cleaned_filename = clean_filename(file.name)
                file.name = cleaned_filename  # Изменяем имя файла

                # Используем update_or_create для обновления или создания записи
                uploaded_file, created = FileUpload.objects.update_or_create(
                    parent=item,  # Условие поиска существующей записи
                    defaults={
                        'file': file,  # Поле файла
                        'uploaded_at': now(),  # Поле даты загрузки
                    }
                )

                # Логируем результат
                if created:
                    logger.info(f"Файл {file.name} успешно создан.")
                else:
                    logger.info(f"Файл {file.name} успешно обновлён.")

            except IntegrityError as e:
                # Логируем ошибку, если что-то пошло не так
                logger.error(f"Ошибка при сохранении файла: {e}")

            return redirect('edit-book', item.id)

    else:
        logger.info("func profile, method GET")
        invoice_list = InvoiceDNRDetails.objects.all()
        form = UploadFileForm()
        context = {
            'invoices': invoice_list,
            'form': form,
            'result': message,
        }
    return render(request, 'registration/profile.html', context=context)


